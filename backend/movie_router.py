"""
电影相关 API 路由
所有路由以 /movie 为前缀
"""

import io
from typing import Optional

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from .db import get_movie_db

router = APIRouter(prefix="/movie", tags=["电影管理"])


class UpdateWatchedDateRequest(BaseModel):
    """更新观影日期请求"""

    watched_date: str


class AddRecordRequest(BaseModel):
    """添加观影记录请求"""

    tmdb_id: int
    watched_date: str
    rating: Optional[float] = None
    review: str = ""


# 获取数据库实例
_db = None


def get_db():
    """获取数据库实例（懒加载）"""
    global _db
    if _db is None:
        _db = get_movie_db()
    return _db


@router.get("/records")
def get_movie_records(
    year: str = Query("", description="观看年份"),
    month: str = Query("", description="观看月份"),
    rating_min: float = Query(None, ge=0, le=10, description="最低评分"),
    rating_max: float = Query(None, ge=0, le=10, description="最高评分"),
    decade: str = Query("", description="出品年代 (如 1990s, 2010s)"),
    vote_count_min: int = Query(None, ge=0, description="最少评分人数"),
    genres: str = Query("", description="电影主题/类型"),
    search: str = Query("", description="搜索电影名称"),
    sort_by: str = Query(
        "watched_date",
        description="排序字段",
        pattern="^(watched_date|vote_average|vote_count|release_date)$",
    ),
    sort_order: str = Query("desc", description="排序方向", pattern="^(asc|desc)$"),
    page: int = Query(1, ge=1, description="页码"),
    limit: int = Query(20, ge=1, le=100, description="每页数量"),
    multiple_watch: bool = Query(False, description="是否只显示多次观看的电影"),
):
    """获取观影记录列表"""
    db = get_db()
    records, total = db.get_records(
        year=year,
        month=month,
        rating_min=rating_min,
        rating_max=rating_max,
        decade=decade,
        vote_count_min=vote_count_min,
        genres=genres,
        search=search,
        sort_by=sort_by,
        sort_order=sort_order,
        page=page,
        limit=limit,
        multiple_watch=multiple_watch,
    )

    return {
        "success": True,
        "data": records,
        "pagination": {
            "page": page,
            "limit": limit,
            "total": total,
            "pages": (total + limit - 1) // limit if total > 0 else 0,
        },
    }


@router.get("/years")
def get_years(
    month: str = Query("", description="观看月份"),
    decade: str = Query("", description="出品年代"),
    rating_min: float = Query(None, ge=0, le=10, description="最低评分"),
    rating_max: float = Query(None, ge=0, le=10, description="最高评分"),
    vote_count_min: int = Query(None, ge=0, description="最少评分人数"),
    search: str = Query("", description="搜索电影名称"),
):
    """获取所有观影年份（联动筛选）"""
    db = get_db()
    years = db.get_years(
        month=month,
        decade=decade,
        rating_min=rating_min,
        rating_max=rating_max,
        vote_count_min=vote_count_min,
        search=search,
    )
    return {"success": True, "data": years}


@router.get("/months")
def get_months(
    year: str = Query("", description="观看年份"),
    decade: str = Query("", description="出品年代"),
    rating_min: float = Query(None, ge=0, le=10, description="最低评分"),
    rating_max: float = Query(None, ge=0, le=10, description="最高评分"),
    vote_count_min: int = Query(None, ge=0, description="最少评分人数"),
    search: str = Query("", description="搜索电影名称"),
):
    """获取所有观影月份（联动筛选）"""
    db = get_db()
    months = db.get_months(
        year=year,
        decade=decade,
        rating_min=rating_min,
        rating_max=rating_max,
        vote_count_min=vote_count_min,
        search=search,
    )
    return {"success": True, "data": months}


@router.get("/decades")
def get_decades(
    year: str = Query("", description="观看年份"),
    month: str = Query("", description="观看月份"),
    rating_min: float = Query(None, ge=0, le=10, description="最低评分"),
    rating_max: float = Query(None, ge=0, le=10, description="最高评分"),
    vote_count_min: int = Query(None, ge=0, description="最少评分人数"),
    search: str = Query("", description="搜索电影名称"),
):
    """获取所有出品年代（联动筛选）"""
    db = get_db()
    decades = db.get_decades(
        year=year,
        month=month,
        rating_min=rating_min,
        rating_max=rating_max,
        vote_count_min=vote_count_min,
        search=search,
    )
    return {"success": True, "data": decades}


@router.get("/stats")
def get_stats():
    """获取观影统计"""
    db = get_db()
    stats = db.get_stats()
    return {
        "success": True,
        "data": stats,
    }


@router.get("/records/{tmdb_id}")
def get_movie_detail(tmdb_id: int):
    """获取电影详情"""
    db = get_db()
    result = db.get_movie_detail(tmdb_id)

    if not result:
        raise HTTPException(status_code=404, detail="电影不存在")

    return {
        "success": True,
        "data": result,
    }


@router.put("/records/{tmdb_id}/date")
def update_watched_date(
    tmdb_id: int,
    old_date: str = Query(..., description="原观影日期"),
    request: UpdateWatchedDateRequest = None,
):
    """更新观影日期"""
    if not request:
        raise HTTPException(status_code=400, detail="请求体不能为空")

    db = get_db()
    success, error_msg = db.update_watched_date(tmdb_id, old_date, request.watched_date)

    if not success:
        if error_msg and "已存在" in error_msg:
            raise HTTPException(status_code=409, detail=error_msg)
        else:
            raise HTTPException(status_code=404, detail=error_msg or "观影记录不存在")

    return {"success": True, "message": "更新成功"}


@router.delete("/records/{tmdb_id}")
def delete_record(tmdb_id: int, watched_date: str = Query(..., description="观影日期")):
    """删除观影记录（不删除电影信息）"""
    db = get_db()
    success = db.delete_record(tmdb_id, watched_date)

    if not success:
        raise HTTPException(status_code=404, detail="观影记录不存在")

    return {"success": True, "message": "删除成功"}


@router.post("/records")
def add_record(request: AddRecordRequest):
    """添加观影记录"""
    db = get_db()

    # 检查电影是否存在
    if not db.movie_exists(request.tmdb_id):
        raise HTTPException(
            status_code=404, detail="电影不存在，请先搜索并添加电影信息"
        )

    success = db.add_record(
        tmdb_id=request.tmdb_id,
        watched_date=request.watched_date,
        rating=request.rating,
        review=request.review,
    )

    if not success:
        raise HTTPException(status_code=500, detail="添加记录失败")

    return {"success": True, "message": "添加成功"}


@router.get("/search")
async def search_movies(
    query: str = Query(..., min_length=1, description="搜索关键词"),
    page: int = Query(1, ge=1, description="页码"),
):
    """搜索TMDB电影"""
    from .tmdb_client import TMDBClient

    client = TMDBClient()
    try:
        results = await client.search_movies(query, page=page)
    finally:
        await client.close()

    return {
        "success": True,
        "data": results,
        "page": page,
    }


@router.get("/movie-detail")
async def get_movie_detail_by_id(
    id: int = Query(..., description="电影ID"),
):
    """按ID获取电影详情"""
    from .tmdb_client import TMDBClient

    client = TMDBClient()
    try:
        movie_data = await client.get_movie_detail(id)
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"无法获取电影信息: {str(e)}")
    finally:
        await client.close()

    if not movie_data:
        raise HTTPException(status_code=404, detail="无法获取电影信息")

    return {
        "success": True,
        "data": movie_data,
    }


@router.post("/movies/{tmdb_id}/add")
async def add_movie_from_tmdb(tmdb_id: int):
    """从TMDB添加电影信息到数据库"""
    from .db.base import MovieInfo
    from .tmdb_client import TMDBClient

    db = get_db()

    # 检查电影是否已存在
    if db.movie_exists(tmdb_id):
        return {"success": True, "message": "电影已存在", "exists": True}

    client = TMDBClient()
    try:
        movie_data = await client.get_movie_detail(tmdb_id)
    except Exception as e:
        await client.close()
        raise HTTPException(status_code=404, detail=f"无法获取电影信息: {str(e)}")
    finally:
        await client.close()

    if not movie_data:
        raise HTTPException(status_code=404, detail="无法获取电影信息")

    movie_info = MovieInfo(
        tmdb_id=movie_data.get("id"),
        title=movie_data.get("title", ""),
        original_title=movie_data.get("original_title", ""),
        overview=movie_data.get("overview", ""),
        release_date=movie_data.get("release_date", ""),
        poster_path=movie_data.get("poster_path", ""),
        backdrop_path=movie_data.get("backdrop_path", ""),
        genres=", ".join([g["name"] for g in movie_data.get("genres", [])]),
        vote_average=movie_data.get("vote_average", 0),
        vote_count=movie_data.get("vote_count", 0),
        runtime=movie_data.get("runtime", 0),
    )

    db.add_movie_info(movie_info)

    return {"success": True, "message": "电影已添加", "data": movie_info.to_dict()}


@router.get("/export")
def export_records(
    year: str = Query("", description="观看年份"),
    month: str = Query("", description="观看月份"),
    rating_min: float = Query(None, ge=0, le=10, description="最低评分"),
    rating_max: float = Query(None, ge=0, le=10, description="最高评分"),
    decade: str = Query("", description="出品年代 (如 1990s, 2010s)"),
    vote_count_min: int = Query(None, ge=0, description="最少评分人数"),
    search: str = Query("", description="搜索电影名称"),
    sort_by: str = Query(
        "watched_date",
        description="排序字段",
        pattern="^(watched_date|vote_average|vote_count|release_date)$",
    ),
    sort_order: str = Query("desc", description="排序方向", pattern="^(asc|desc)$"),
):
    """导出观影记录为Excel"""
    db = get_db()

    # 获取所有符合条件的记录（不分页）
    records, total = db.get_records(
        year=year,
        month=month,
        rating_min=rating_min,
        rating_max=rating_max,
        decade=decade,
        vote_count_min=vote_count_min,
        search=search,
        sort_by=sort_by,
        sort_order=sort_order,
        page=1,
        limit=10000,  # 足够大以获取所有记录
    )

    # 创建Excel
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "观影记录"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="4A5568", end_color="4A5568", fill_type="solid"
    )
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写入表头
    headers = [
        "电影名称",
        "原名",
        "观看日期",
        "上映日期",
        "类型",
        "TMDB评分",
        "评分人数",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, record in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=record.get("title", "")).border = (
            thin_border
        )
        ws.cell(
            row=row_idx, column=2, value=record.get("original_title", "")
        ).border = thin_border
        ws.cell(row=row_idx, column=3, value=record.get("watched_date", "")).border = (
            thin_border
        )
        ws.cell(row=row_idx, column=4, value=record.get("release_date", "")).border = (
            thin_border
        )
        ws.cell(row=row_idx, column=5, value=record.get("genres", "")).border = (
            thin_border
        )
        ws.cell(row=row_idx, column=6, value=record.get("vote_average", 0)).border = (
            thin_border
        )
        ws.cell(row=row_idx, column=7, value=record.get("vote_count", 0)).border = (
            thin_border
        )

    # 调整列宽
    column_widths = [30, 30, 15, 15, 20, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # 生成文件名
    from urllib.parse import quote

    filename = "观影记录"
    if year:
        filename += f"_{year}年"
    if month:
        filename += f"_{month}月"
    filename += ".xlsx"
    # URL编码处理中文文件名
    encoded_filename = quote(filename)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        },
    )


@router.get("/statistics-yearly")
def get_statistics_yearly():
    """获取历年观影统计数据（单独端点，不受年份过滤影响）"""
    db = get_db()
    with db._get_connection() as conn:
        cursor = conn.cursor()

        # 按年份统计（排除NULL值）
        cursor.execute(
            """
            SELECT
                strftime('%Y', replace(r.watched_date, '/', '-')) as year,
                COUNT(DISTINCT r.id) as count
            FROM movie_records r
            WHERE r.watched_date IS NOT NULL AND r.watched_date != ''
            GROUP BY year
            ORDER BY year DESC
        """
        )
        by_year = [{"year": row[0], "count": row[1]} for row in cursor.fetchall()]

    return {
        "success": True,
        "data": by_year,
    }


@router.get("/statistics")
def get_statistics(
    year: str = Query("", description="按观看年份筛选，格式：YYYY"),
):
    """获取观影统计数据（按出品年代、评分、主题，支持年份过滤）"""
    db = get_db()
    with db._get_connection() as conn:
        cursor = conn.cursor()

        # 构建年份过滤条件
        year_filter = ""
        params = []
        if year:
            year_filter = "AND strftime('%Y', replace(r.watched_date, '/', '-')) = ?"
            params = [year]
        cursor.execute(
            f"""
            SELECT
                CASE
                    WHEN substr(m.release_date, 1, 4) >= '2020' THEN '2020s'
                    WHEN substr(m.release_date, 1, 4) >= '2010' THEN '2010s'
                    WHEN substr(m.release_date, 1, 4) >= '2000' THEN '2000s'
                    WHEN substr(m.release_date, 1, 4) >= '1990' THEN '1990s'
                    WHEN substr(m.release_date, 1, 4) >= '1980' THEN '1980s'
                    WHEN substr(m.release_date, 1, 4) >= '1970' THEN '1970s'
                    WHEN substr(m.release_date, 1, 4) >= '1960' THEN '1960s'
                    WHEN substr(m.release_date, 1, 4) >= '1950' THEN '1950s'
                    ELSE '1940s及更早'
                END as decade,
                COUNT(DISTINCT r.id) as count
            FROM movie_records r
            INNER JOIN movie_info m ON m.tmdb_id = r.tmdb_id
            WHERE m.release_date IS NOT NULL AND m.release_date != '' {year_filter}
            GROUP BY decade
            ORDER BY decade DESC
        """,
            params,
        )
        by_decade = [{"decade": row[0], "count": row[1]} for row in cursor.fetchall()]

        # 按评分范围统计
        cursor.execute(
            f"""
            SELECT
                CASE
                    WHEN m.vote_average >= 9 THEN '9-10分'
                    WHEN m.vote_average >= 8 THEN '8-9分'
                    WHEN m.vote_average >= 7 THEN '7-8分'
                    WHEN m.vote_average >= 6 THEN '6-7分'
                    ELSE '低于6分'
                END as range,
                COUNT(DISTINCT r.id) as count
            FROM movie_records r
            INNER JOIN movie_info m ON m.tmdb_id = r.tmdb_id {year_filter}
            GROUP BY range
            ORDER BY range DESC
        """,
            params,
        )
        by_rating = [{"range": row[0], "count": row[1]} for row in cursor.fetchall()]

        # 按主题统计（按数量从多到少排序）
        cursor.execute(
            f"""
            SELECT m.genres
            FROM movie_records r
            INNER JOIN movie_info m ON m.tmdb_id = r.tmdb_id
            WHERE m.genres IS NOT NULL AND m.genres != '' {year_filter}
            """,
            params,
        )
        rows = cursor.fetchall()
        genres_count = {}
        for row in rows:
            genres_str = row[0]
            if genres_str:
                # genres 存储为"类型1, 类型2"的格式
                genres_list = [g.strip() for g in genres_str.split(",")]
                for genre in genres_list:
                    genres_count[genre] = genres_count.get(genre, 0) + 1

        # 按数量从多到少排序
        by_genre = [
            {"genre": genre, "count": count}
            for genre, count in sorted(
                genres_count.items(), key=lambda x: x[1], reverse=True
            )
        ]

    return {
        "success": True,
        "data": {
            "by_decade": by_decade,
            "by_rating": by_rating,
            "by_genre": by_genre,
        },
    }


@router.get("/genres")
def get_genres(
    year: str = Query("", description="观看年份"),
    month: str = Query("", description="观看月份"),
    decade: str = Query("", description="出品年代"),
    rating_min: float = Query(None, ge=0, le=10, description="最低评分"),
    rating_max: float = Query(None, ge=0, le=10, description="最高评分"),
    vote_count_min: int = Query(None, ge=0, description="最少评分人数"),
    search: str = Query("", description="搜索电影名称"),
):
    """获取所有电影主题/类型列表（根据当前过滤条件）"""
    db = get_db()
    genres = db.get_genres(
        year=year,
        month=month,
        decade=decade,
        rating_min=rating_min,
        rating_max=rating_max,
        vote_count_min=vote_count_min,
        search=search,
    )

    return {
        "success": True,
        "data": genres,
    }


@router.get("/records-with-null-dates")
def get_records_with_null_dates():
    """获取那些 watched_date 为 null 或为空的观影记录"""
    db = get_db()
    with db._get_connection() as conn:
        cursor = conn.cursor()
        cursor.execute(
            """
            SELECT
                m.tmdb_id,
                m.title,
                m.original_title,
                r.watched_date
            FROM movie_records r
            INNER JOIN movie_info m ON m.tmdb_id = r.tmdb_id
            WHERE r.watched_date IS NULL OR r.watched_date = ''
            ORDER BY m.title
        """
        )
        records = [
            {
                "tmdb_id": row[0],
                "title": row[1],
                "original_title": row[2],
                "watched_date": row[3],
            }
            for row in cursor.fetchall()
        ]

    return {
        "success": True,
        "data": records,
    }


@router.post("/refresh-from-tmdb")
def refresh_movies_from_tmdb():
    """查询所有电影的最新数据，返回需要更新的字段"""
    db = get_db()

    try:
        # 获取所有电影
        with db._get_connection() as conn:
            cursor = conn.cursor()
            cursor.execute(
                """
                SELECT
                    m.tmdb_id,
                    m.title,
                    m.vote_average,
                    m.vote_count,
                    m.release_date,
                    m.overview
                FROM movie_info m
                ORDER BY m.title
            """
            )
            all_movies = []
            for row in cursor.fetchall():
                all_movies.append(
                    {
                        "tmdb_id": row[0],
                        "title": row[1],
                        "current": {
                            "vote_average": row[2],
                            "vote_count": row[3],
                            "release_date": row[4],
                            "overview": row[5],
                        },
                    }
                )

        # 这里应该调用TMDB API获取最新数据并对比
        # 暂时返回空列表，表示无需更新
        changes = []

        return {"success": True, "data": {"changes": changes}}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"刷新电影数据失败: {str(e)}")
